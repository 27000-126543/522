from typing import Optional, List, Dict, Any
from datetime import datetime, timedelta
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from app.models.models import (
    DailyReport, Turbine, WorkOrder, WindFarm, Warning,
    OrderStatus, ProcessingRecord, FaultHistory
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class ReportService:
    @staticmethod
    def generate_daily_report(
        db: Session,
        report_date: Optional[datetime] = None,
        wind_farm_id: Optional[int] = None
    ) -> List[DailyReport]:
        if report_date is None:
            report_date = datetime.now()
        start_of_day = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_of_day = start_of_day + timedelta(days=1)

        reports = []

        turbines_query = db.query(Turbine)
        if wind_farm_id:
            turbines_query = turbines_query.filter(Turbine.wind_farm_id == wind_farm_id)
        all_turbines = turbines_query.all()

        farm_model_groups: Dict[tuple, List[Turbine]] = {}
        for t in all_turbines:
            key = (t.wind_farm_id, t.model or "Unknown")
            if key not in farm_model_groups:
                farm_model_groups[key] = []
            farm_model_groups[key].append(t)

        for (wf_id, model), turbines in farm_model_groups.items():
            turbine_ids = [t.id for t in turbines]

            work_orders = db.query(WorkOrder).filter(
                WorkOrder.turbine_id.in_(turbine_ids),
                WorkOrder.created_at >= start_of_day,
                WorkOrder.created_at < end_of_day
            ).all()

            fault_count = len(work_orders)
            total_turbines = len(turbines)
            fault_rate = round((fault_count / total_turbines * 100) if total_turbines > 0 else 0, 2)

            completed_orders = [wo for wo in work_orders if wo.status == OrderStatus.COMPLETED]
            total_repair_hours = 0.0
            for wo in completed_orders:
                if wo.started_at and wo.completed_at:
                    total_repair_hours += (wo.completed_at - wo.started_at).total_seconds() / 3600
            avg_repair_hours = round(
                total_repair_hours / len(completed_orders), 2
            ) if completed_orders else 0.0

            total_downtime = 0.0
            for wo in work_orders:
                if wo.accepted_at and wo.completed_at:
                    total_downtime += (wo.completed_at - wo.accepted_at).total_seconds() / 3600

            spare_parts_consumed: Dict[str, int] = {}
            for wo in work_orders:
                records = db.query(ProcessingRecord).filter(
                    ProcessingRecord.work_order_id == wo.id
                ).all()
                for rec in records:
                    if rec.spare_parts:
                        for sp in rec.spare_parts if isinstance(rec.spare_parts, list) else []:
                            if isinstance(sp, dict):
                                name = sp.get("name", sp.get("part_code", "Unknown"))
                                qty = int(sp.get("quantity", 0))
                                spare_parts_consumed[name] = spare_parts_consumed.get(name, 0) + qty

            warnings_count = db.query(Warning).filter(
                Warning.turbine_id.in_(turbine_ids),
                Warning.timestamp >= start_of_day,
                Warning.timestamp < end_of_day
            ).count()

            report = DailyReport(
                report_date=start_of_day,
                wind_farm_id=wf_id,
                turbine_model=model,
                total_turbines=total_turbines,
                fault_count=fault_count,
                fault_rate=fault_rate,
                avg_repair_hours=avg_repair_hours,
                total_downtime_hours=round(total_downtime, 2),
                spare_parts_consumed=spare_parts_consumed if spare_parts_consumed else None,
                warnings_count={"total": warnings_count},
                work_orders_completed=len(completed_orders),
                work_orders_pending=len([wo for wo in work_orders
                                         if wo.status not in [OrderStatus.COMPLETED, OrderStatus.CANCELLED]])
            )
            db.add(report)
            reports.append(report)

        db.flush()
        return reports

    @staticmethod
    def query_reports(
        db: Session,
        start_date: datetime,
        end_date: Optional[datetime] = None,
        wind_farm_id: Optional[int] = None,
        turbine_model: Optional[str] = None
    ) -> List[DailyReport]:
        query = db.query(DailyReport).filter(DailyReport.report_date >= start_date)
        if end_date:
            query = query.filter(DailyReport.report_date <= end_date)
        if wind_farm_id:
            query = query.filter(DailyReport.wind_farm_id == wind_farm_id)
        if turbine_model:
            query = query.filter(DailyReport.turbine_model == turbine_model)
        return query.order_by(DailyReport.report_date.desc()).all()

    @staticmethod
    def export_to_excel(
        db: Session,
        reports: List[DailyReport],
        start_date: datetime,
        end_date: datetime
    ) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = "运维报表"

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.merge_cells('A1:K1')
        ws['A1'] = f"智慧风电运维报表 ({start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')})"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws['A1'].alignment = center_align

        headers = [
            "日期", "风电场", "风机型号", "总台数",
            "故障数", "故障率(%)", "平均修复时长(小时)",
            "总停机时长(小时)", "工单完成数", "工单数", "备件消耗"
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        wind_farm_cache = {}
        for idx, report in enumerate(reports, 3):
            if report.wind_farm_id and report.wind_farm_id not in wind_farm_cache:
                wf = db.query(WindFarm).filter(WindFarm.id == report.wind_farm_id).first()
                wind_farm_cache[report.wind_farm_id] = wf.name if wf else f"场#{report.wind_farm_id}"
            wf_name = wind_farm_cache.get(report.wind_farm_id, "-")

            spare_parts_str = "-"
            if report.spare_parts_consumed:
                parts = [f"{k}×{v}" for k, v in report.spare_parts_consumed.items()]
                spare_parts_str = "; ".join(parts)

            row_data = [
                report.report_date.strftime("%Y-%m-%d"),
                wf_name,
                report.turbine_model or "-",
                report.total_turbines,
                report.fault_count,
                report.fault_rate,
                report.avg_repair_hours,
                report.total_downtime_hours,
                report.work_orders_completed,
                report.work_orders_pending,
                spare_parts_str
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=idx, column=col_idx, value=value)
                cell.alignment = center_align
                cell.border = thin_border

        col_widths = [12, 18, 18, 8, 8, 10, 16, 16, 12, 10, 40]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

        summary_row = len(reports) + 3
        ws.cell(row=summary_row, column=1, value="汇总").font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value="=SUM(D3:D" + str(summary_row - 1) + ")").font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value="=SUM(E3:E" + str(summary_row - 1) + ")").font = Font(bold=True)
        ws.cell(row=summary_row, column=8, value="=SUM(H3:H" + str(summary_row - 1) + ")").font = Font(bold=True)
        ws.cell(row=summary_row, column=9, value="=SUM(I3:I" + str(summary_row - 1) + ")").font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def get_statistics_summary(
        db: Session,
        wind_farm_id: Optional[int] = None,
        days: int = 30
    ) -> Dict[str, Any]:
        start_date = datetime.now() - timedelta(days=days)

        turbine_q = db.query(Turbine)
        if wind_farm_id:
            turbine_q = turbine_q.filter(Turbine.wind_farm_id == wind_farm_id)
        total_turbines = turbine_q.count()

        wo_q = db.query(WorkOrder).filter(WorkOrder.created_at >= start_date)
        if wind_farm_id:
            turbine_ids = [t.id for t in turbine_q.all()]
            wo_q = wo_q.filter(WorkOrder.turbine_id.in_(turbine_ids))

        work_orders = wo_q.all()
        total_faults = len(work_orders)
        completed_count = len([wo for wo in work_orders if wo.status == OrderStatus.COMPLETED])
        pending_count = len([wo for wo in work_orders if wo.status not in [OrderStatus.COMPLETED, OrderStatus.CANCELLED]])

        avg_repair = 0.0
        repair_times = []
        for wo in work_orders:
            if wo.started_at and wo.completed_at:
                repair_times.append((wo.completed_at - wo.started_at).total_seconds() / 3600)
        if repair_times:
            avg_repair = round(sum(repair_times) / len(repair_times), 2)

        persistent_count = db.query(Turbine).filter(
            Turbine.is_persistent_risk == True
        ).count()
        if wind_farm_id:
            persistent_count = db.query(Turbine).filter(
                Turbine.wind_farm_id == wind_farm_id,
                Turbine.is_persistent_risk == True
            ).count()

        return {
            "period_days": days,
            "total_turbines": total_turbines,
            "total_faults": total_faults,
            "fault_rate": round((total_faults / total_turbines * 100) if total_turbines > 0 else 0, 2),
            "completed_orders": completed_count,
            "pending_orders": pending_count,
            "avg_repair_hours": avg_repair,
            "persistent_risk_count": persistent_count,
            "healthy_turbines": total_turbines - persistent_count
        }
